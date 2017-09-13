# -*- coding: utf-8 -*-
# requed:
# https://openpyxl.readthedocs.io/en/default/
# 
# require:
# pip install pyyaml
# pip install polib
# 
import argparse
import os, os.path
import xml.etree.ElementTree as ET
from openpyxl import Workbook, load_workbook
from scripts import *
import polib

def xml2excel(xmlpath, excelpath):
    print(xmlpath)
    removefile(excelpath)

    xml = ET.parse(xmlpath)
    xml = xml.getroot()
    row = len(xml)
    if row == 0:
        print("can not export empty xml to excel " + xmlpath)
        return

    wb = Workbook()
    ws = wb.active
    oneRows = xml[0]
    col = len(oneRows)
    # print(len(xmls), xmlpath)
    # l = len(xmls)
    # xmls[0]
    # row = len(cols)
    for j in xrange(1, col+1):
        # print(oneRows[j - 1].tag)
        # print(ws.cell(1, j))
        ws.cell(row = 1, column = j, value = oneRows[j - 1].tag)
    for i in xrange(2, row + 2):
        oneRows = xml[i - 2]
        for j in xrange(1, col + 1):
            ws.cell(row = i, column = j, value = oneRows[j - 1].text)

    wb.save(os.path.abspath(excelpath))


def exportI18n(originPath, i18nPath):
    print(originPath)
    wb = load_workbook(os.path.abspath(originPath))
    ws = wb.active
    col = ws.max_column
    row = ws.max_row
    if row <= 1:
        print('!!!!!empty excel!!!!!')
        return
    # print(col, row)
    # ws.Columns.AutoFit()
    
    titles = readrow(ws, 1)
    
    # for i in xrange(1, row):
        

    xml = ET.Element(os.path.basename(xmlpath)[0:-4])
    xml.set('xmlns:xsi', 'http://www.w3.org/2001/XMLSchema-instance')
    for i in xrange(2, row + 1):
        table = ET.SubElement(xml, 'table')
        for j in range(1, col + 1):
            one = ET.SubElement(table, titles[j - 1])
            one.text = ws.cell(row = i, column = j).value
    indent(xml)
    data = ET.tostring(xml, 'utf-8')
    writefile(xmlpath, xmldeclaration + data)

def importI18ns(xmldir, exceldir):
    makedirs(exceldir)
    for root, dirs, files in os.walk(xmldir):
        for f in files:
            if f.endswith('.xml'):
                xmlpath = os.path.join(root, f)
                relpath = os.path.relpath(xmlpath, xmldir)
                excelpath = os.path.join(exceldir, relpath[:-3] + 'xlsx')
                xml2excel(xmlpath, excelpath)

def mergepo(po, kmap):
    for i in xrange(1,len(po) + 1):
        entry = po[i]
        if entry.msgid in kmap:
            kmap[entry] = False
    for k, v in kmap.iteritems():
        entry = polib.POEntry(
            msgid=k,
            msgstr=u'',
            obsolete = False
        )
        po.append(entry)
    # for key in kmap.iteritems():
    #     entry = po.find(key, include_obsolete_entries = True)

def exportI18ns(originDir, i18nDir, cfgmap, adapter):
    makedirs(i18nDir)
    kmap = {}
    for filename, cols in cfgmap.iteritems():
        filename = os.path.join(originDir, filename)
        adapter.readfile(filename, cols, kmap)

    pofilename = os.path.join(i18nDir, 'zh_TW.po')
    po = loadpo(pofilename)
    mergepo(po, kmap)
    po.save(pofilename)

    # for root, dirs, files in os.walk(originDir):
    #     for f in files:
    #         if f.endswith('.xlsx'):
    #             originPath = os.path.join(root, f)
    #             relpath = os.path.relpath(originPath, i18nDir)
    #             i18nPath = os.path.join(xmldir, relpath)
    #             exportI18n(originPath, i18nPath)


# -------------- main ----------------
if __name__ == '__main__':
    # parser = argparse.ArgumentParser(usage='%(prog)s <method> <xmldir> <exceldir>',
    #     description='.xml <==> .xlsx\nexport xml  files to xlsx\nexport xlsx files to xml',
    #     formatter_class=argparse.RawTextHelpFormatter)
    # parser.add_argument('method', choices=['x2e', 'e2x'],
    #     help='x2e: xml to excel\ne2x: excel to xml')
    # parser.add_argument('xmldir',
    #     help='xml files directory')
    # parser.add_argument('exceldir',
    #     help='excel files directory')

    # args = parser.parse_args()

    # if args.method == 'x2e':
    #     xml2excels(args.xmldir, args.exceldir)
    # elif args.method == 'e2x':
    #     excel2xmls(args.exceldir, args.xmldir)
    #     
    #     
    # excelpath = 'zh_code.xlsx'
    # wb = load_workbook(os.path.abspath(excelpath))
    # ws = wb.active
    # col = ws.max_column
    # row = ws.max_row
    # if row <= 1:
    #     print('!!!!!empty excel!!!!!', col, row)
        
    # # print(col, row)
    # # ws.Columns.AutoFit()
    # cell = ws.cell(row = 1, column = 1)
    # value = cell.value
    # print cell.value
    # comment = cell.comment.text
    # # comment.encoding
    # print(comment)
    # print cell.comment.text.startswith(ur'多语言')
    # print ur'多语言'
    # writefile('xxxxx.txt', 'this is text')
    # 
    # 
    pass
cfgmap = readyaml('config.yml')

import scripts.adapters.csvadapter as csvadapter
exportI18ns('csv', 'languages', cfgmap, csvadapter)
