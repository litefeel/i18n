# -*- coding: utf-8 -*-
# 
import os, os.path
from openpyxl import Workbook, load_workbook
from ... import isascii

def gettitles(ws, rowIdx = 1, genDict = False):
    max_col = ws.max_column
    max_row = ws.max_row
    titles = {} if genDict else []
    i = 1
    for row in ws.iter_rows(min_row=rowIdx, max_col=max_col, max_row=rowIdx):
        for cell in row:
            if genDict:
                titles[cell.value] = i
                i += 1
            else:
                titles.append(cell.value or '')
    return titles


def readfile(filename, colNames, kmap):
    wb = load_workbook(os.path.abspath(filename))
    ws = wb.active
    max_col = ws.max_column
    max_row = ws.max_row

    titlemap = gettitles(ws, 2, True)
    for colName in colNames:
        colidx = titlemap[colName]
        for col in ws.iter_cols(min_row=4, max_row=max_row, min_col=colidx, max_col=colidx):
            for cell in col:
                v = cell.value
                if v:
                    kmap[v] = True


# wite file1 to file2
def writefile(filename1, filename2, colNames, kvmap):
    wb = load_workbook(os.path.abspath(filename1))
    ws = wb.active
    max_col = ws.max_column
    max_row = ws.max_row

    titlemap = gettitles(ws, 2, True)
    for colName in colNames:
        colidx = titlemap[colName]
        for col in ws.iter_cols(min_row=4, max_row=max_row, min_col=colidx, max_col=colidx):
            for cell in col:
                v = cell.value
                if v and v in kvmap:
                    cell.value = kvmap[v]

    wb.save(os.path.abspath(filename2))

def checkcols(filename):
    map = {}
    cols = []
    wb = load_workbook(os.path.abspath(filename))
    ws = wb.active
    max_col = ws.max_column
    max_row = ws.max_row

    titles = gettitles(ws, 2)
    i = 0
    for col in ws.iter_cols(min_row=4, max_col=max_col, max_row=max_row):
        i += 1
        key = titles[i - 1]
        if not key or key in map:
            continue
        for cell in col:
            v = str(cell.value) if cell.value else None
            if v and not isascii(v):
                map[key] = True
                cols.append(key)
                break
    return cols

