# -*- coding: utf-8 -*-
# requed:
# https://openpyxl.readthedocs.io/en/default/
# 
import os, os.path
from openpyxl import Workbook, load_workbook
from ...function import loadpo
from ...function import makedirs
from ...function import makedirs2


def read(langsdir, lang):
    filename = os.path.join(langsdir, lang + '.xlsx')
    if not os.path.exists(filename):
        return {}

    kvmap = {}
    wb = load_workbook(os.path.abspath(filename))
    ws = wb.active
    col = ws.max_column
    max_row = ws.max_row
    for row in xrange(1, max_row + 1):
        k = ws.cell(row = row, column = 1).value
        v = ws.cell(row = row, column = 2).value or ''
        if k:
            kvmap[k] = v

    return kvmap

def save(langsdir, lang, kmap):
    filename = os.path.join(langsdir, lang + '.xlsx')
    kvmap = {}
    for k, v in kmap.iteritems():
        kvmap[k] = ''
    oldmap = read(langsdir, lang)
    if oldmap:
        for k, v in oldmap.iteritems():
            if k in kvmap:
                kvmap[k] = v
    wb = Workbook()
    ws = wb.active

    row = 1
    for k, v in kvmap.iteritems():
        ws.cell(row = row, column = 1, value = k)
        ws.cell(row = row, column = 2, value = v)
        row += 1
    
    wb.save(os.path.abspath(filename))
